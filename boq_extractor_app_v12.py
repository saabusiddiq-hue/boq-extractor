"""
BOQ Extractor Pro v14.0 - Dark Theme Edition
PDF to Excel/CSV Converter with Auto-Preview & Interactive Crop
"""

import streamlit as st
import pandas as pd
import numpy as np
import io
import re
import base64
from datetime import datetime
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from PIL import Image, ImageDraw, ImageFont
import fitz  # PyMuPDF

# Page configuration - DARK THEME
st.set_page_config(
    page_title="BOQ Extractor Pro",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded"
)

# DARK THEME CUSTOM CSS
def load_css():
    st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
        
        /* Main app background - Dark */
        .stApp {
            background-color: #0f1117 !important;
        }
        
        /* Sidebar - Darker */
        [data-testid="stSidebar"] {
            background-color: #161b22 !important;
            border-right: 1px solid #30363d;
        }
        
        /* Headers */
        h1, h2, h3, h4, h5, h6 {
            color: #e6edf3 !important;
            font-family: 'Inter', sans-serif;
            font-weight: 600;
        }
        
        /* Main header styling */
        .main-header {
            background: linear-gradient(135deg, #1f6feb 0%, #388bfd 100%);
            color: white;
            padding: 2rem;
            border-radius: 16px;
            text-align: center;
            margin-bottom: 2rem;
            box-shadow: 0 10px 40px rgba(31, 111, 235, 0.3);
            border: 1px solid #30363d;
        }
        
        .main-header h1 {
            font-size: 2.5rem;
            font-weight: 700;
            margin-bottom: 0.5rem;
            color: white !important;
        }
        
        .main-header p {
            font-size: 1.1rem;
            opacity: 0.9;
            margin-bottom: 1rem;
            color: #e6edf3;
        }
        
        /* Badge styling */
        .badge-container {
            display: flex;
            gap: 10px;
            justify-content: center;
            flex-wrap: wrap;
        }
        
        .feature-badge {
            background-color: rgba(255,255,255,0.15);
            color: white;
            padding: 6px 16px;
            border-radius: 20px;
            font-size: 0.85rem;
            font-weight: 500;
            backdrop-filter: blur(10px);
            border: 1px solid rgba(255,255,255,0.2);
        }
        
        /* Section headers */
        .section-header {
            background: linear-gradient(90deg, #21262d 0%, #161b22 100%);
            padding: 1rem 1.5rem;
            border-radius: 10px;
            margin: 2rem 0 1rem 0;
            border-left: 4px solid #1f6feb;
            border: 1px solid #30363d;
        }
        
        .section-header h3 {
            margin: 0;
            color: #58a6ff !important;
            font-weight: 600;
        }
        
        /* Cards */
        .stat-card {
            background: #161b22;
            border-radius: 12px;
            padding: 1.5rem;
            box-shadow: 0 4px 20px rgba(0,0,0,0.3);
            border-left: 4px solid #1f6feb;
            border: 1px solid #30363d;
        }
        
        .stat-value {
            font-size: 2rem;
            font-weight: 700;
            color: #58a6ff;
        }
        
        .stat-label {
            font-size: 0.9rem;
            color: #8b949e;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        
        /* Preview box */
        .preview-box {
            border: 3px solid #1f6feb;
            border-radius: 12px;
            overflow: hidden;
            position: relative;
            background: #0d1117;
        }
        
        /* Extracted text area */
        .extracted-text {
            background: #0d1117;
            color: #e6edf3;
            padding: 1.5rem;
            border-radius: 12px;
            font-family: 'Courier New', monospace;
            font-size: 0.9rem;
            max-height: 400px;
            overflow-y: auto;
            white-space: pre-wrap;
            border: 1px solid #30363d;
        }
        
        /* Slider labels */
        .slider-label {
            font-weight: 600;
            color: #e6edf3;
            margin-bottom: 0.5rem;
            font-size: 0.9rem;
        }
        
        /* Crop info box */
        .crop-info {
            background: #21262d;
            border-radius: 8px;
            padding: 1rem;
            border: 1px solid #30363d;
            margin-top: 1rem;
        }
        
        .crop-info h4 {
            color: #58a6ff !important;
            margin-bottom: 0.5rem;
        }
        
        .crop-dimension {
            display: flex;
            justify-content: space-between;
            padding: 0.3rem 0;
            border-bottom: 1px solid #30363d;
            color: #c9d1d9;
        }
        
        /* Buttons */
        .stButton>button {
            background: linear-gradient(135deg, #238636 0%, #2ea043 100%) !important;
            color: white !important;
            border: none !important;
            border-radius: 8px !important;
            font-weight: 600 !important;
            transition: all 0.2s !important;
        }
        
        .stButton>button:hover {
            transform: translateY(-2px) !important;
            box-shadow: 0 4px 12px rgba(35, 134, 54, 0.4) !important;
        }
        
        .stButton>button[kind="secondary"] {
            background: #21262d !important;
            border: 1px solid #30363d !important;
            color: #c9d1d9 !important;
        }
        
        /* Input fields */
        .stTextInput>div>div>input, .stNumberInput>div>div>input {
            background-color: #0d1117 !important;
            color: #e6edf3 !important;
            border: 1px solid #30363d !important;
            border-radius: 6px !important;
        }
        
        /* Sliders */
        .stSlider>div>div>div {
            background-color: #1f6feb !important;
        }
        
        /* File uploader */
        .stFileUploader>div>div {
            background-color: #161b22 !important;
            border: 2px dashed #30363d !important;
            border-radius: 10px !important;
            color: #8b949e !important;
        }
        
        /* Data editor */
        .stDataFrame {
            background-color: #161b22 !important;
            border: 1px solid #30363d !important;
            border-radius: 10px !important;
        }
        
        /* Success/Warning/Error messages */
        .stSuccess {
            background-color: rgba(35, 134, 54, 0.1) !important;
            border: 1px solid #238636 !important;
            color: #3fb950 !important;
        }
        
        .stWarning {
            background-color: rgba(210, 153, 34, 0.1) !important;
            border: 1px solid #d29922 !important;
            color: #e3b341 !important;
        }
        
        .stError {
            background-color: rgba(248, 81, 73, 0.1) !important;
            border: 1px solid #f85149 !important;
            color: #f85149 !important;
        }
        
        /* Info boxes */
        .stInfo {
            background-color: rgba(56, 139, 253, 0.1) !important;
            border: 1px solid #388bfd !important;
            color: #58a6ff !important;
        }
        
        /* Custom scrollbar */
        ::-webkit-scrollbar {
            width: 8px;
            height: 8px;
        }
        
        ::-webkit-scrollbar-track {
            background: #161b22;
            border-radius: 4px;
        }
        
        ::-webkit-scrollbar-thumb {
            background: #1f6feb;
            border-radius: 4px;
        }
        
        /* Footer */
        .footer {
            text-align: center;
            padding: 2rem;
            color: #8b949e;
            border-top: 1px solid #30363d;
            margin-top: 3rem;
        }
        
        /* Text colors */
        p, span, label {
            color: #c9d1d9 !important;
        }
    </style>
    """, unsafe_allow_html=True)

# =============================================================================
# SMART BOQ PARSER
# =============================================================================

class BOQParser:
    """Intelligent BOQ parser using pattern recognition"""
    
    def __init__(self):
        self.material_patterns = [
            r'A240\s+SS316', r'A240\s+SS304', r'A36\b', r'A105\b',
            r'A193\s+GR\.B7', r'A194\s+GR\.2H', r'A193\s+B7', r'A194\s+2H',
            r'Per\s+MSS-SP[0-9]+', r'CARBON\s+STEEL', r'SS316', r'SS304',
            r'A516\s+GR\.60', r'A516\s+GR\.70', r'A240\b', r'A193\b',
            r'A194\b', r'A516\b', r'Graphite\s+bronze', r'BRONZE', r'SS\s+316'
        ]
        
        self.part_patterns = [
            r'^[A-Z][0-9]+-[A-Z][0-9]+$',
            r'^V[0-9]+-[0-9]+-[A-Z]+[0-9]*$',
            r'^M[0-9]+x[0-9]+$',
            r'^PC[0-9]+-[0-9\-]+$',
            r'^[0-9]+x[0-9]+x[0-9]+$',
            r'^[A-Z][0-9]+$',
            r'^[0-9]+-[A-Z][0-9]+$',
        ]

    def is_item_number(self, text):
        try:
            num = int(text.strip())
            return 1 <= num <= 999
        except:
            return False

    def is_quantity(self, text):
        try:
            num = int(text.strip())
            return 1 <= num <= 9999
        except:
            return False

    def is_part_number(self, text):
        text = text.strip()
        if not text:
            return False
        return any(re.match(pattern, text, re.IGNORECASE) for pattern in self.part_patterns)

    def extract_material(self, text):
        text = text.strip()
        for pattern in self.material_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return match.group(0), text[:match.start()].strip(), text[match.end():].strip()
        return "", text, ""

    def parse_line(self, line):
        line = line.strip()
        if not line or len(line) < 10:
            return None

        skip_keywords = ['ITEM', 'NO.', 'QTY', 'QUANTITY', 'PART', 'DESCRIPTION', 
                        'MATERIAL', 'DRAWING', 'TITLE', 'NOTE:', 'DATE:']
        first_word = line.split()[0].upper() if line.split() else ""
        if any(kw in first_word for kw in skip_keywords):
            return None

        parts = line.split()
        if len(parts) < 4:
            return None

        item_no = None
        quantity = None
        part_no = None
        description = ""
        material = ""

        if self.is_item_number(parts[0]) and self.is_quantity(parts[1]):
            item_no = int(parts[0])
            quantity = int(parts[1])
            remaining = parts[2:]
            
            part_idx = -1
            for i, part in enumerate(remaining[:3]):
                if self.is_part_number(part):
                    part_no = part
                    part_idx = i
                    break
            
            if part_no is None:
                for i, part in enumerate(remaining[:3]):
                    if len(part) >= 3 and any(c.isdigit() for c in part) and any(c.isalpha() for c in part):
                        part_no = part
                        part_idx = i
                        break
            
            if part_no:
                remaining_str = ' '.join(remaining[part_idx+1:])
                mat, desc_before, desc_after = self.extract_material(remaining_str)
                if mat:
                    material = mat
                    description = desc_before
                else:
                    description = remaining_str
            else:
                description = ' '.join(remaining)
        else:
            return None

        if item_no is None or quantity is None:
            return None

        return {
            'Item No': item_no,
            'Quantity': quantity,
            'Part No': part_no or "",
            'Description': description.strip(),
            'Material': material
        }

    def parse_text_block(self, text):
        items = []
        lines = text.split('\n')
        for line in lines:
            parsed = self.parse_line(line)
            if parsed:
                items.append(parsed)
        return items

# =============================================================================
# PDF PROCESSING FUNCTIONS
# =============================================================================

def render_pdf_page(pdf_file, page_num, crop_box=None, zoom_level=2.0):
    """Render PDF page with optional crop box overlay"""
    try:
        pdf_file.seek(0)
        doc = fitz.open(stream=pdf_file.read(), filetype="pdf")
        
        if page_num > len(doc):
            doc.close()
            return None, None, "Invalid page number"
        
        page = doc[page_num - 1]
        rect = page.rect
        width_pt, height_pt = rect.width, rect.height
        
        # Convert points to pixels (72 DPI standard, scaled by zoom)
        width_px = int(width_pt * zoom_level)
        height_px = int(height_pt * zoom_level)
        
        # Calculate crop coordinates
        crop_coords_px = None
        if crop_box:
            x1_pct, y1_pct, x2_pct, y2_pct = crop_box
            x1 = width_pt * (x1_pct / 100)
            y1 = height_pt * (y1_pct / 100)
            x2 = width_pt * (x2_pct / 100)
            y2 = height_pt * (y2_pct / 100)
            
            # Pixel coordinates for display
            crop_coords_px = {
                'x': int(x1 * zoom_level),
                'y': int(y1 * zoom_level),
                'width': int((x2 - x1) * zoom_level),
                'height': int((y2 - y1) * zoom_level),
                'x1_pct': x1_pct,
                'y1_pct': y1_pct,
                'x2_pct': x2_pct,
                'y2_pct': y2_pct
            }
            
            # Draw crop rectangle
            crop_rect = fitz.Rect(x1, y1, x2, y2)
            page.draw_rect(crop_rect, color=(0.97, 0.32, 0.29), width=3)
            
            # Semi-transparent fill using shape
            shape = page.new_shape()
            shape.draw_rect(crop_rect)
            shape.finish(color=(0.97, 0.32, 0.29), fill=(0.97, 0.32, 0.29), fill_opacity=0.15)
            shape.commit()
            
            # Add label
            label_point = fitz.Point(x1 + 5, y1 - 5)
            page.insert_text(label_point, "EXTRACTION AREA", 
                           fontsize=14, color=(0.97, 0.32, 0.29), fontname="helv")
        
        # Render
        mat = fitz.Matrix(zoom_level, zoom_level)
        pix = page.get_pixmap(matrix=mat)
        img_data = pix.tobytes("png")
        
        doc.close()
        return img_data, crop_coords_px, None
        
    except Exception as e:
        return None, None, str(e)

def extract_text_from_region(pdf_file, page_num, crop_box):
    """Extract text from specific region of PDF"""
    try:
        pdf_file.seek(0)
        with pdfplumber.open(pdf_file) as pdf:
            page = pdf.pages[page_num - 1]
            width, height = page.width, page.height
            
            x1_pct, y1_pct, x2_pct, y2_pct = crop_box
            x1 = width * (x1_pct / 100)
            y1 = height * (y1_pct / 100)
            x2 = width * (x2_pct / 100)
            y2 = height * (y2_pct / 100)
            
            cropped = page.crop((x1, y1, x2, y2))
            text = cropped.extract_text()
            return text
            
    except Exception as e:
        return f"Error extracting text: {str(e)}"

def process_pdf_batch(pdf_file, start_page, end_page, crop_box, progress_callback=None):
    """Process multiple pages with the same crop region"""
    all_items = []
    logs = []
    
    try:
        pdf_file.seek(0)
        with pdfplumber.open(pdf_file) as pdf:
            total_pages = len(pdf.pages)
            actual_end = min(end_page, total_pages) if end_page != 999 else total_pages
            
            logs.append(f"📁 Processing pages {start_page} to {actual_end} (Total: {total_pages})")
            
            parser = BOQParser()
            
            for page_num in range(start_page - 1, actual_end):
                if progress_callback:
                    progress = (page_num - start_page + 2) / (actual_end - start_page + 1)
                    progress_callback(min(progress, 1.0), f"Processing page {page_num + 1}...")
                
                page = pdf.pages[page_num]
                width, height = page.width, page.height
                
                x1_pct, y1_pct, x2_pct, y2_pct = crop_box
                x1 = width * (x1_pct / 100)
                y1 = height * (y1_pct / 100)
                x2 = width * (x2_pct / 100)
                y2 = height * (y2_pct / 100)
                
                cropped = page.crop((x1, y1, x2, y2))
                text = cropped.extract_text()
                
                items = parser.parse_text_block(text)
                
                full_text = page.extract_text()
                drawing_no = ""
                
                dwg_patterns = [
                    r'Q\d+\s+SUPP\s+\d+-\d+',
                    r'DWG[.\s]*[A-Z0-9\-]+',
                    r'DRAWING[.\s]*NO[.\s]*[A-Z0-9\-]+'
                ]
                for pattern in dwg_patterns:
                    match = re.search(pattern, full_text, re.IGNORECASE)
                    if match:
                        drawing_no = match.group(0)
                        break
                
                for item in items:
                    item['Drawing No'] = drawing_no
                    item['Page'] = page_num + 1
                
                status = "✅" if items else "⚠️"
                logs.append(f"{status} Page {page_num+1}: {len(items)} items extracted")
                all_items.extend(items)
                
    except Exception as e:
        logs.append(f"❌ Error: {str(e)}")
    
    return all_items, logs

# =============================================================================
# EXCEL EXPORT - FIXED VERSION
# =============================================================================

def create_excel_download(df, yellow_header=True):
    """Create styled Excel file - FIXED column iteration"""
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ Data"
    
    # Styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True, size=11, color="000000")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Reorder columns
    preferred_order = ['Drawing No', 'Page', 'Item No', 'Quantity', 'Part No', 'Description', 'Material']
    available_cols = [c for c in preferred_order if c in df.columns]
    other_cols = [c for c in df.columns if c not in preferred_order]
    final_cols = available_cols + other_cols
    
    df_export = df[final_cols]
    
    # Write headers
    for col_num, header in enumerate(df_export.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        if yellow_header:
            cell.fill = yellow_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Write data
    for row_num, row_data in enumerate(df_export.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    
    # FIXED: Auto-adjust column widths - iterate properly through rows
    for idx, col in enumerate(df_export.columns, 1):
        max_length = len(str(col))
        
        # Get column letter
        if idx <= 26:
            col_letter = chr(64 + idx)
        else:
            col_letter = chr(64 + (idx-1)//26) + chr(65 + (idx-1)%26)
        
        # Iterate through all rows to find max length
        for row_idx in range(1, len(df_export) + 2):
            cell_value = ws.cell(row=row_idx, column=idx).value
            try:
                cell_length = len(str(cell_value)) if cell_value is not None else 0
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Freeze header
    ws.freeze_panes = 'A2'
    
    wb.save(output)
    output.seek(0)
    return output

# =============================================================================
# MAIN APPLICATION
# =============================================================================

def main():
    load_css()
    
    # Header
    st.markdown("""
    <div class="main-header">
        <h1>📋 BOQ Extractor Pro v14.0</h1>
        <p>Dark Edition - Auto Preview & Smart Extraction</p>
        <div class="badge-container">
            <span class="feature-badge">🎯 Auto Preview</span>
            <span class="feature-badge">📐 Live Crop Adjust</span>
            <span class="feature-badge">📊 Excel Export</span>
            <span class="feature-badge">🎨 Dark UI</span>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Initialize session state
    if 'extracted_data' not in st.session_state:
        st.session_state.extracted_data = pd.DataFrame()
    if 'extraction_logs' not in st.session_state:
        st.session_state.extraction_logs = []
    if 'crop_coords' not in st.session_state:
        st.session_state.crop_coords = (10, 20, 90, 80)
    if 'preview_text' not in st.session_state:
        st.session_state.preview_text = ""
    if 'zoom_level' not in st.session_state:
        st.session_state.zoom_level = 1.5
    if 'pdf_file' not in st.session_state:
        st.session_state.pdf_file = None
    if 'auto_preview' not in st.session_state:
        st.session_state.auto_preview = False
    
    # Sidebar controls
    with st.sidebar:
        st.markdown("### 📐 Region Configuration")
        
        st.markdown("<div class='slider-label'>Left (X1) %</div>", unsafe_allow_html=True)
        x1 = st.slider("X1", 0, 100, st.session_state.crop_coords[0], key="x1_slider", label_visibility="collapsed")
        
        st.markdown("<div class='slider-label'>Top (Y1) %</div>", unsafe_allow_html=True)
        y1 = st.slider("Y1", 0, 100, st.session_state.crop_coords[1], key="y1_slider", label_visibility="collapsed")
        
        st.markdown("<div class='slider-label'>Right (X2) %</div>", unsafe_allow_html=True)
        x2 = st.slider("X2", 0, 100, st.session_state.crop_coords[2], key="x2_slider", label_visibility="collapsed")
        
        st.markdown("<div class='slider-label'>Bottom (Y2) %</div>", unsafe_allow_html=True)
        y2 = st.slider("Y2", 0, 100, st.session_state.crop_coords[3], key="y2_slider", label_visibility="collapsed")
        
        # Ensure valid coordinates
        if x2 <= x1:
            x2 = min(x1 + 10, 100)
        if y2 <= y1:
            y2 = min(y1 + 10, 100)
            
        current_crop = (x1, y1, x2, y2)
        st.session_state.crop_coords = current_crop
        
        # Calculate crop dimensions
        crop_width_pct = x2 - x1
        crop_height_pct = y2 - y1
        
        st.markdown("---")
        
        # Crop Size Display
        st.markdown("### 📏 Crop Size")
        with st.container():
            st.markdown(f"""
            <div class="crop-info">
                <h4>Dimensions</h4>
                <div class="crop-dimension">
                    <span>Width:</span>
                    <span>{crop_width_pct}% of page</span>
                </div>
                <div class="crop-dimension">
                    <span>Height:</span>
                    <span>{crop_height_pct}% of page</span>
                </div>
                <div class="crop-dimension" style="border-bottom: none;">
                    <span>Area:</span>
                    <span>{crop_width_pct * crop_height_pct / 100:.1f}% of page</span>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        st.markdown("### 🔍 Preview Settings")
        zoom_level = st.select_slider(
            "Zoom Level",
            options=[1.0, 1.5, 2.0, 2.5, 3.0],
            value=st.session_state.zoom_level,
            format_func=lambda x: f"{x}x"
        )
        st.session_state.zoom_level = zoom_level
        
        st.markdown("---")
        
        st.markdown("### 📄 Page Settings")
        preview_page = st.number_input("Preview Page", 1, 1000, 1, key="preview_page")
        start_page = st.number_input("Start Page", 1, 1000, 1, key="start_page")
        end_page = st.number_input("End Page (999=all)", 1, 1000, 999, key="end_page")
        
        st.markdown("---")
        
        st.markdown("### 🎨 Export Options")
        yellow_header = st.checkbox("Yellow Excel Header", value=True)
        
        st.markdown("---")
        
        if st.button("🗑️ Clear All Data", type="secondary"):
            st.session_state.extracted_data = pd.DataFrame()
            st.session_state.extraction_logs = []
            st.session_state.preview_text = ""
            st.session_state.pdf_file = None
            st.session_state.auto_preview = False
            st.rerun()
    
    # Main content area
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown("<div class='section-header'><h3>📤 Upload PDF</h3></div>", unsafe_allow_html=True)
        
        uploaded_file = st.file_uploader(
            "Choose a PDF file containing BOQ tables",
            type=['pdf'],
            help="Upload a PDF with Bill of Quantities. Preview will generate automatically."
        )
        
        if uploaded_file:
            st.success(f"✅ Loaded: {uploaded_file.name} ({uploaded_file.size:,} bytes)")
            st.session_state.pdf_file = uploaded_file
            st.session_state.auto_preview = True
    
    with col2:
        st.markdown("<div class='section-header'><h3>🚀 Actions</h3></div>", unsafe_allow_html=True)
        
        if uploaded_file:
            extract_btn = st.button("🔍 Extract BOQ", use_container_width=True, type="primary")
    
    # AUTO-PREVIEW SECTION - Updates automatically when sliders change
    if st.session_state.pdf_file and st.session_state.auto_preview:
        st.markdown("<div class='section-header'><h3>🎯 Live Preview</h3></div>", unsafe_allow_html=True)
        
        # Check if crop coordinates changed to trigger re-render
        if 'last_crop' not in st.session_state:
            st.session_state.last_crop = current_crop
            should_render = True
        else:
            should_render = st.session_state.last_crop != current_crop
            if should_render:
                st.session_state.last_crop = current_crop
        
        # Render preview
        with st.spinner("Updating preview..."):
            img_data, crop_coords_px, error = render_pdf_page(
                st.session_state.pdf_file, 
                preview_page, 
                current_crop, 
                zoom_level
            )
            
            if error:
                st.error(f"Error: {error}")
            else:
                # Display image with full width
                st.image(img_data, caption=f"Page {preview_page} - Red box shows extraction area (Zoom: {zoom_level}x)", use_column_width=True)
                
                # Show pixel dimensions if available
                if crop_coords_px:
                    st.markdown(f"""
                    <div style="display: flex; gap: 2rem; margin-top: 0.5rem; color: #8b949e; font-size: 0.9rem;">
                        <span>📐 Crop Area: {crop_coords_px['width']} × {crop_coords_px['height']} px</span>
                        <span>📍 Position: ({crop_coords_px['x']}, {crop_coords_px['y']}) px</span>
                    </div>
                    """, unsafe_allow_html=True)
                
                # Two columns for details and sample text
                col_details, col_text = st.columns([1, 1])
                
                with col_details:
                    st.markdown("### 📊 Crop Coordinates")
                    st.info(f"""
                    **Current Selection:**
                    - Left (X1): {x1}%
                    - Top (Y1): {y1}%
                    - Right (X2): {x2}%
                    - Bottom (Y2): {y2}%
                    
                    **Dimensions:**
                    - Width: {crop_width_pct}% of page
                    - Height: {crop_height_pct}% of page
                    
                    **Tips:**
                    - Adjust sliders to cover the BOQ table
                    - Red box shows what will be extracted
                    - Higher zoom = more precise selection
                    """)
                
                with col_text:
                    # Extract and show sample text
                    sample_text = extract_text_from_region(st.session_state.pdf_file, preview_page, current_crop)
                    st.session_state.preview_text = sample_text
                    
                    if sample_text:
                        st.markdown("### 📝 Sample Extracted Text")
                        st.markdown(f"<div class='extracted-text'>{sample_text[:1500]}{'...' if len(sample_text) > 1500 else ''}</div>", unsafe_allow_html=True)
                        
                        # Parse sample
                        parser = BOQParser()
                        sample_items = parser.parse_text_block(sample_text)
                        
                        if sample_items:
                            st.success(f"✅ Found {len(sample_items)} items in preview")
                        else:
                            st.warning("⚠️ No BOQ items detected. Adjust region or check format.")
                    else:
                        st.warning("No text found in selected region")
    
    # EXTRACTION SECTION
    if uploaded_file and extract_btn:
        st.markdown("<div class='section-header'><h3>⚙️ Processing</h3></div>", unsafe_allow_html=True)
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        def update_progress(progress, message):
            progress_bar.progress(progress, message)
        
        with st.spinner("Extracting BOQ data..."):
            items, logs = process_pdf_batch(uploaded_file, start_page, end_page, current_crop, update_progress)
            
            st.session_state.extraction_logs = logs
            
            if items:
                df = pd.DataFrame(items)
                st.session_state.extracted_data = df
                progress_bar.empty()
                status_text.success(f"✅ Successfully extracted {len(df)} items!")
            else:
                progress_bar.empty()
                status_text.error("❌ No items extracted. Check region settings.")
    
    # RESULTS SECTION
    if not st.session_state.extracted_data.empty:
        df = st.session_state.extracted_data
        
        st.markdown("<div class='section-header'><h3>📊 Extraction Results</h3></div>", unsafe_allow_html=True)
        
        # Statistics
        c1, c2, c3, c4 = st.columns(4)
        
        with c1:
            st.markdown(f"""
            <div class="stat-card">
                <div class="stat-value">{len(df)}</div>
                <div class="stat-label">Total Items</div>
            </div>
            """, unsafe_allow_html=True)
        
        with c2:
            unique_drawings = df['Drawing No'].nunique() if 'Drawing No' in df.columns else 0
            st.markdown(f"""
            <div class="stat-card">
                <div class="stat-value">{unique_drawings}</div>
                <div class="stat-label">Drawings</div>
            </div>
            """, unsafe_allow_html=True)
        
        with c3:
            unique_materials = df['Material'].nunique() if 'Material' in df.columns else 0
            st.markdown(f"""
            <div class="stat-card">
                <div class="stat-value">{unique_materials}</div>
                <div class="stat-label">Materials</div>
            </div>
            """, unsafe_allow_html=True)
        
        with c4:
            total_qty = int(df['Quantity'].sum()) if 'Quantity' in df.columns else 0
            st.markdown(f"""
            <div class="stat-card">
                <div class="stat-value">{total_qty:,}</div>
                <div class="stat-label">Total Quantity</div>
            </div>
            """, unsafe_allow_html=True)
        
        # Data Editor
        st.markdown("### ✏️ Review and Edit Data")
        
        edited_df = st.data_editor(
            df,
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            column_config={
                "Drawing No": st.column_config.TextColumn("Drawing No", width="medium"),
                "Page": st.column_config.NumberColumn("Page", width="small"),
                "Item No": st.column_config.NumberColumn("Item No", width="small"),
                "Quantity": st.column_config.NumberColumn("Qty", width="small"),
                "Part No": st.column_config.TextColumn("Part No", width="medium"),
                "Description": st.column_config.TextColumn("Description", width="large"),
                "Material": st.column_config.TextColumn("Material", width="medium"),
            }
        )
        
        # Export options
        st.markdown("<div class='section-header'><h3>📦 Export Data</h3></div>", unsafe_allow_html=True)
        
        col_excel, col_csv = st.columns(2)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        with col_excel:
            try:
                excel_data = create_excel_download(edited_df, yellow_header)
                st.download_button(
                    label="📥 Download Excel (.xlsx)",
                    data=excel_data,
                    file_name=f"BOQ_Extract_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as e:
                st.error(f"Excel export error: {str(e)}")
        
        with col_csv:
            csv_data = edited_df.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="📄 Download CSV (.csv)",
                data=csv_data,
                file_name=f"BOQ_Extract_{timestamp}.csv",
                mime="text/csv",
                use_container_width=True
            )
    
    # Logs
    if st.session_state.extraction_logs:
        with st.expander("🔍 Processing Logs", expanded=False):
            for log in st.session_state.extraction_logs:
                if log.startswith("✅"):
                    st.markdown(f"<span style='color: #3fb950;'>{log}</span>", unsafe_allow_html=True)
                elif log.startswith("❌"):
                    st.markdown(f"<span style='color: #f85149;'>{log}</span>", unsafe_allow_html=True)
                elif log.startswith("⚠️"):
                    st.markdown(f"<span style='color: #d29922;'>{log}</span>", unsafe_allow_html=True)
                else:
                    st.markdown(f"<span style='color: #58a6ff;'>{log}</span>", unsafe_allow_html=True)
    
    # Instructions
    with st.expander("📖 How to Use", expanded=False):
        st.markdown("""
        ### Quick Start Guide

        **1. Upload PDF** 📤
        - Upload your BOQ PDF file
        - Preview generates **automatically**
        
        **2. Adjust Crop Region** 📐
        - Use X1, Y1, X2, Y2 sliders in sidebar
        - Watch preview update **live**
        - Check "Crop Size" for dimensions
        
        **3. Fine-tune Zoom** 🔍
        - Select zoom level (1x - 3x)
        - Higher zoom = more precise selection
        
        **4. Extract Data** 🔍
        - Click "Extract BOQ" button
        - Review extracted items
        
        **5. Export** 📦
        - Download as Excel (yellow headers) or CSV
        - Edit data in the table before export if needed
        
        ### Supported Formats
        - **Item No**: 1, 2, 3...
        - **Quantity**: Integer numbers
        - **Part No**: M1-V1, V2-26-BM1, etc.
        - **Material**: A240 SS316, A36, etc.
        """)
    
    # Footer
    st.markdown("""
    <div class="footer">
        <p>📋 BOQ Extractor Pro v14.0 | Dark Edition</p>
        <p style="font-size: 0.8rem; color: #8b949e;">
            Auto-preview | Live crop adjust | Excel export
        </p>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
