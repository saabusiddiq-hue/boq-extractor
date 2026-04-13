import streamlit as st
import pandas as pd
import numpy as np
import io
import re
import shutil
from datetime import datetime
from typing import List, Dict, Tuple, Optional

# Setup
TESSERACT_PATH = shutil.which("tesseract")
OCR_AVAILABLE = False
if TESSERACT_PATH:
    try:
        import pytesseract
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
        OCR_AVAILABLE = True
    except ImportError:
        pass

try:
    import pdfplumber
except ImportError:
    st.error("pdfplumber not found")
    st.stop()

try:
    import fitz
except ImportError:
    st.error("PyMuPDF not found")
    st.stop()

from PIL import Image, ImageEnhance, ImageFilter

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("openpyxl not found")
    st.stop()

st.set_page_config(page_title="BOQ Extractor Pro v28", page_icon="📋", layout="wide")

st.markdown("""
<style>
    .stApp { background-color: #0d1117; }
    .upload-bar { background: #161b22; border: 1px solid #30363d; border-radius: 8px; padding: 1rem; margin-bottom: 1rem; }
    .left-panel { background: #161b22; border: 2px solid #30363d; border-radius: 8px; padding: 1rem; }
    .right-panel { background: #161b22; border: 1px solid #30363d; border-radius: 8px; padding: 1rem; min-height: 600px; }
    .extract-btn { background: linear-gradient(90deg, #238636 0%, #2ea043 100%) !important; color: white !important; font-weight: 600 !important; padding: 1rem !important; font-size: 1.1rem !important; }
    .scan-btn { background: linear-gradient(90deg, #1f6feb 0%, #58a6ff 100%) !important; color: white !important; font-weight: 600 !important; }
    .info-box { background: rgba(31, 111, 235, 0.1); border: 1px solid #1f6feb; border-radius: 6px; padding: 0.75rem; margin: 0.5rem 0; font-size: 0.85rem; }
    .fig-no-tag { background: rgba(35, 134, 54, 0.2); color: #3fb950; padding: 2px 6px; border-radius: 4px; font-weight: 500; }
    .column-detected { background: rgba(46, 160, 67, 0.2); border: 1px solid #3fb950; border-radius: 4px; padding: 4px 8px; margin: 2px; display: inline-block; font-size: 0.8rem; }
    .anchor-tag { background: rgba(245, 158, 11, 0.2); color: #f59e0b; padding: 2px 6px; border-radius: 4px; font-weight: 500; font-size: 0.8rem; }
</style>
""", unsafe_allow_html=True)


class ColumnScanner:
    """Scan and detect column positions automatically with manual anchor support"""
    def __init__(self):
        self.columns = []
        self.positions = {}  # column_name -> (start, end) positions
        self.header_line = ""
        self.material_anchors = []  # Manual material anchor points

    def set_material_anchors(self, anchors: List[str]):
        """Set manual material anchor keywords"""
        self.material_anchors = [a.strip().upper() for a in anchors if a.strip()]

    def scan_pdf_page(self, pdf_bytes: bytes, page_num: int = 0, crop: Tuple = None, 
                      manual_anchors: List[str] = None) -> Dict:
        """Scan a PDF page to detect column structure with optional manual anchors"""

        if manual_anchors:
            self.set_material_anchors(manual_anchors)

        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            if page_num >= len(pdf.pages):
                page_num = 0
            page = pdf.pages[page_num]

            if crop:
                w, h = page.width, page.height
                page = page.crop((w*crop[0]/100, h*crop[1]/100, w*crop[2]/100, h*crop[3]/100))

            # Get words with positions
            words = page.extract_words()

            # Try to find header row
            text = page.extract_text() or ""
            lines = text.split("\n")

            # Look for header patterns
            header_patterns = [
                r"Item\s+No\.\s+Req'd\s+Fig\.\s+No\.",
                r"Item\s+No\.\s+Qty\s+Fig",
                r"ITEM\s+QTY\s+FIG",
                r"No\.\s+Req'd\s+Fig",
            ]

            detected_columns = []
            header_found = False

            for line in lines[:15]:  # Check first 15 lines
                line_stripped = line.strip()
                if not line_stripped:
                    continue

                # Check for header patterns
                line_upper = line_stripped.upper()

                # Detect column headers based on keywords
                if any(keyword in line_upper for keyword in ["ITEM", "NO.", "REQ'D", "QTY", "FIG", "DESCRIPTION", "MATERIAL"]):
                    self.header_line = line_stripped

                    # Find positions of each keyword
                    words_in_line = line_stripped.split()
                    current_pos = 0

                    for word in words_in_line:
                        word_upper = word.upper()
                        # Categorize words into columns
                        if word_upper in ["ITEM", "NO.", "No."] and not any(c["name"] == "Item No" for c in detected_columns):
                            detected_columns.append({"name": "Item No", "keyword": word, "position": current_pos})
                        elif word_upper in ["REQ'D", "QTY", "QUANTITY"] and not any(c["name"] == "Qty" for c in detected_columns):
                            detected_columns.append({"name": "Qty", "keyword": word, "position": current_pos})
                        elif word_upper in ["FIG.", "FIG", "FIG.NO.", "FIG_NO", "PART"] and not any(c["name"] == "Fig No" for c in detected_columns):
                            detected_columns.append({"name": "Fig No", "keyword": word, "position": current_pos})
                        elif word_upper in ["DESCRIPTION", "DESC"] and not any(c["name"] == "Description" for c in detected_columns):
                            detected_columns.append({"name": "Description", "keyword": word, "position": current_pos})
                        elif word_upper in ["MATERIAL", "MAT'L", "GRADE"] and not any(c["name"] == "Material" for c in detected_columns):
                            detected_columns.append({"name": "Material", "keyword": word, "position": current_pos})
                        current_pos += len(word) + 1

                    if len(detected_columns) >= 3:
                        header_found = True
                        break

            # If no header found, try pattern matching on data rows
            if not header_found:
                for line in lines[5:30]:  # Check data rows
                    if re.match(r"^\s*\d+\s+\d+", line):  # Starts with two numbers
                        # This looks like a data row, infer structure
                        parts = line.split()
                        if len(parts) >= 4:
                            detected_columns = [
                                {"name": "Item No", "position": 0},
                                {"name": "Qty", "position": 1},
                                {"name": "Fig No", "position": 2},
                                {"name": "Description", "position": 3},
                            ]
                            if len(parts) > 4:
                                detected_columns.append({"name": "Material", "position": 4})
                            break

            self.columns = detected_columns
            return {
                "header_found": header_found,
                "columns": detected_columns,
                "header_line": self.header_line,
                "sample_lines": lines[:20],
                "material_anchors_used": self.material_anchors
            }

    def get_column_structure(self) -> List[str]:
        """Return list of column names in order"""
        return [c["name"] for c in self.columns]


class BOQExtractor:
    """Extract BOQ with detected column structure and manual material anchors"""
    def __init__(self, column_structure: List[str] = None, material_anchors: List[str] = None):
        self.column_structure = column_structure or ["Item No", "Qty", "Fig No", "Description", "Material"]
        self.material_anchors = material_anchors or []
        # Default material patterns + user anchors
        self.material_patterns = [
            r"(A36|A105|A193|A194|A240|A516)\b",
            r"(SS316|SS316L|SS304|SS304L)\b",
            r"(Per\s+MSS\-SP\d+|MSS\-SP\d+)",
            r"(Gr\.\s*\d+\.?\d*)",
            r"(CI\.\s*\d+|Cast\s+Iron)",
            r"(Bronze|Graphite|PTFE|Carbon\s+Steel)",
            r"(GR\.\s*\d+[A-Z]?)",
        ]
        # Add custom anchors as patterns
        for anchor in self.material_anchors:
            if anchor:
                # Escape special regex characters
                safe_anchor = re.escape(anchor)
                self.material_patterns.append(f"({safe_anchor})\b")

    def extract_line(self, line: str) -> Optional[Dict]:
        """Extract data from a single line based on column structure"""
        line = line.strip()
        if not line or not line[0].isdigit():
            return None

        # Split by multiple spaces/tabs to handle blank columns
        parts = re.split(r"\s{2,}|\t", line)
        parts = [p.strip() for p in parts if p.strip()]

        if len(parts) < 3:
            # Try single space split if multi-space fails
            parts = line.split()

        result = {}

        try:
            # Item No - first number
            result["Item"] = int(parts[0])

            # Qty - second number
            if len(parts) > 1 and parts[1].isdigit():
                result["Qty"] = int(parts[1])
                idx = 2
            else:
                result["Qty"] = 1
                idx = 1

            # Fig No - can be compound (e.g., "Graphite bronze")
            fig_no_parts = []
            if idx < len(parts):
                fig_no_parts.append(parts[idx])
                idx += 1

                # Check for compound Fig No (e.g., "Graphite bronze", "SS Plate")
                if idx < len(parts):
                    next_word = parts[idx]
                    if next_word.lower() in ["bronze", "plate", "steel", "pad", "sheet", "rod", "bush", "no.", "no"]:
                        if idx + 1 < len(parts):  # Make sure there's more after
                            fig_no_parts.append(next_word)
                            idx += 1

            result["Fig No"] = " ".join(fig_no_parts) if fig_no_parts else ""

            # Remaining parts are Description and Material
            remaining = parts[idx:] if idx < len(parts) else []

            if remaining:
                full_text = " ".join(remaining)

                # Try to extract Material from end using patterns + anchors
                material_found = ""
                for pattern in self.material_patterns:
                    matches = list(re.finditer(pattern, full_text, re.IGNORECASE))
                    if matches:
                        last_match = matches[-1]
                        remaining_after = full_text[last_match.end():].strip()
                        if len(remaining_after) < 5:
                            material_found = last_match.group(1)
                            description = full_text[:last_match.start()].strip(" -:/")
                            result["Description"] = description
                            result["Material"] = material_found
                            break

                if not material_found:
                    result["Description"] = full_text
                    result["Material"] = ""
            else:
                result["Description"] = ""
                result["Material"] = ""

        except (ValueError, IndexError) as e:
            return None

        return result

    def extract_from_pdf(self, pdf_bytes: bytes, crop: Tuple, max_items: int, page_range: str = "all") -> List[Dict]:
        """Extract BOQ from PDF"""
        items = []

        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            total_pages = len(pdf.pages)

            # Determine pages to process
            if page_range == "all":
                pages_to_process = range(total_pages)
            else:
                try:
                    if "-" in page_range:
                        start, end = map(int, page_range.split("-"))
                        pages_to_process = range(start-1, min(end, total_pages))
                    else:
                        page = int(page_range) - 1
                        pages_to_process = [page] if 0 <= page < total_pages else range(total_pages)
                except:
                    pages_to_process = range(total_pages)

            for page_num in pages_to_process:
                page = pdf.pages[page_num]

                if crop:
                    w, h = page.width, page.height
                    page = page.crop((w*crop[0]/100, h*crop[1]/100, w*crop[2]/100, h*crop[3]/100))

                text = page.extract_text() or ""
                lines = text.split("\n")

                for line in lines:
                    item = self.extract_line(line)
                    if item and item["Item"] <= max_items:
                        item["Page"] = page_num + 1
                        items.append(item)

        return items


@st.cache_data
def render_preview(pdf_bytes, page_num, crop, zoom):
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        if page_num > len(doc):
            return None
        page = doc[page_num - 1]

        if crop and len(crop) == 4:
            x1, y1, x2, y2 = crop
            rect = page.rect
            x1 = rect.width * x1 / 100
            y1 = rect.height * y1 / 100
            x2 = rect.width * x2 / 100
            y2 = rect.height * y2 / 100

            shape = page.new_shape()
            shape.draw_rect(fitz.Rect(x1, y1, x2, y2))
            shape.finish(color=(1, 0, 0), fill=(1, 0, 0), fill_opacity=0.1, width=2)
            shape.commit()
            page.insert_text(fitz.Point(x1 + 5, max(y1 - 5, 10)), "CROP", fontsize=10, color=(1, 0, 0))

        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
        return pix.tobytes("png")
    except:
        return None


def create_excel(df, yellow_header=True):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True, size=10)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))

    cols = ["Item", "Qty", "Fig No", "Description", "Material", "Page"]
    df = df[[c for c in cols if c in df.columns]]

    for col_num, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        if yellow_header:
            cell.fill = yellow_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row_num, row in enumerate(df.values, 2):
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(idx)
        max_len = max([len(str(ws.cell(row=r, column=idx).value or "")) for r in range(1, len(df)+2)])
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    ws.freeze_panes = "A2"
    wb.save(output)
    return output.getvalue()


# Session init
if "scanner" not in st.session_state:
    st.session_state.scanner = ColumnScanner()
if "column_structure" not in st.session_state:
    st.session_state.column_structure = None
if "target_bytes" not in st.session_state:
    st.session_state.target_bytes = None
if "sample_bytes" not in st.session_state:
    st.session_state.sample_bytes = None
if "data" not in st.session_state:
    st.session_state.data = None
if "scan_results" not in st.session_state:
    st.session_state.scan_results = None
if "material_anchors" not in st.session_state:
    st.session_state.material_anchors = []

st.title("📋 BOQ Extractor Pro v28 - Auto Column Detection + Material Anchors")

# TOP BAR - Upload Section
st.markdown("<div class='upload-bar'>", unsafe_allow_html=True)

# Create tabs for Target and Sample upload
tab1, tab2 = st.tabs(["📄 Target PDF", "📑 Sample BOQ (Optional)"])

with tab1:
    col1, col2, col3 = st.columns([2, 2, 3])

    with col1:
        st.write("📄 Upload Target PDF *")
        target_file = st.file_uploader("Target", type=["pdf"], label_visibility="collapsed", key="target_upload")
        if target_file:
            st.session_state.target_bytes = target_file.read()
            st.success("✅ PDF loaded")

    with col2:
        st.write("🔍 Auto-Detect Columns")
        if st.session_state.target_bytes:
            if st.button("SCAN PDF STRUCTURE", use_container_width=True, type="primary"):
                with st.spinner("Scanning column structure..."):
                    scanner = ColumnScanner()
                    # Pass manual anchors if set
                    manual_anchors = st.session_state.get("material_anchors", [])
                    results = scanner.scan_pdf_page(
                        st.session_state.target_bytes, 
                        crop=(5, 15, 95, 60),
                        manual_anchors=manual_anchors
                    )
                    st.session_state.scan_results = results
                    st.session_state.scanner = scanner
                    st.session_state.column_structure = scanner.get_column_structure()

                    if results["header_found"]:
                        st.success(f"✅ Found {len(results['columns'])} columns!")
                    else:
                        st.warning("⚠️ No header found, using default structure")
        else:
            st.info("Upload PDF first")

    with col3:
        if st.session_state.scan_results:
            st.write("📊 Detected Structure:")
            cols = st.session_state.scan_results["columns"]
            for col in cols:
                st.markdown(f"<span class='column-detected'>{col['name']}</span>", unsafe_allow_html=True)
            if st.session_state.scan_results["header_line"]:
                st.caption(f"Header: {st.session_state.scan_results['header_line'][:50]}...")

            # Show if material anchors were used
            if st.session_state.scan_results.get("material_anchors_used"):
                st.markdown("<small>🔧 Material anchors active</small>", unsafe_allow_html=True)

with tab2:
    st.write("📑 Upload Sample BOQ (for reference or template)")
    sample_file = st.file_uploader("Sample BOQ", type=["pdf", "png", "jpg", "jpeg"], 
                                    label_visibility="collapsed", key="sample_upload")
    if sample_file:
        st.session_state.sample_bytes = sample_file.read()
        st.success("✅ Sample loaded for reference")

        # Preview sample if it's an image
        file_type = sample_file.name.split('.')[-1].lower()
        if file_type in ['png', 'jpg', 'jpeg']:
            st.image(st.session_state.sample_bytes, caption="Sample BOQ Preview", use_column_width=True)

st.markdown("</div>", unsafe_allow_html=True)

# MATERIAL ANCHORS SECTION
st.markdown("<div class='upload-bar' style='margin-top: 0.5rem;'>", unsafe_allow_html=True)
anchor_col1, anchor_col2 = st.columns([2, 3])

with anchor_col1:
    st.write("⚓ Material Anchor Points (Optional)")
    st.caption("Add keywords to help detect material columns (e.g., 'ASTM A105', 'SS304', 'Ductile Iron')")

    # Text area for material anchors
    anchors_input = st.text_area(
        "Material Anchors (one per line)",
        value="\n".join(st.session_state.material_anchors) if st.session_state.material_anchors else "",
        placeholder="ASTM A105\nSS304\nDuctile Iron\nCarbon Steel",
        height=100,
        key="material_anchors_input"
    )

    if st.button("💾 Save Anchors", type="secondary"):
        anchors_list = [a.strip() for a in anchors_input.split("\n") if a.strip()]
        st.session_state.material_anchors = anchors_list
        st.success(f"✅ Saved {len(anchors_list)} material anchors!")

with anchor_col2:
    if st.session_state.material_anchors:
        st.write("📌 Active Material Anchors:")
        anchor_cols = st.columns(4)
        for idx, anchor in enumerate(st.session_state.material_anchors):
            with anchor_cols[idx % 4]:
                st.markdown(f"<span class='anchor-tag'>{anchor}</span>", unsafe_allow_html=True)
    else:
        st.info("💡 Material anchors help the extractor identify material specifications in descriptions.\n\nExample: If your BOQ uses 'ASTM A105' as material spec, add it here.")

st.markdown("</div>", unsafe_allow_html=True)

# Show sample data if scan performed
if st.session_state.scan_results and st.session_state.scan_results.get("sample_lines"):
    with st.expander("📄 View Sample Lines from PDF", expanded=False):
        for i, line in enumerate(st.session_state.scan_results["sample_lines"][:10]):
            if line.strip():
                st.text(f"{i+1}: {line}")

# MAIN LAYOUT
left_col, right_col = st.columns([1, 2.5])

# LEFT - Preview & Settings
with left_col:
    st.markdown("<div class='left-panel'>", unsafe_allow_html=True)
    st.write("👁️ Preview & Crop Settings")

    # Page range selection
    page_range = st.text_input("Page Range", value="all", 
                               help="Use 'all', '1-3', or single page like '2'")

    st.divider()

    x = st.slider("X (Left)", 0, 100, 5)
    y = st.slider("Y (Top)", 0, 100, 15)
    z = st.slider("Z (Right)", 0, 100, 95)
    w = st.slider("W (Bottom)", 0, 100, 60)

    if z <= x: z = min(x + 10, 100)
    if w <= y: w = min(y + 10, 100)

    crop = (x, y, z, w)

    # Max items
    max_items = st.number_input("Max Items", min_value=5, max_value=1000, value=100)

    if st.session_state.target_bytes:
        img = render_preview(st.session_state.target_bytes, 1, crop, 2.0)
        if img:
            st.image(img, use_column_width=True)
    else:
        st.info("Upload PDF to preview")

    st.markdown("</div>", unsafe_allow_html=True)

# RIGHT - Extraction & Table
with right_col:
    st.markdown("<div class='right-panel'>", unsafe_allow_html=True)

    if st.session_state.target_bytes:
        col_extract, col_manual = st.columns([1, 1])

        with col_extract:
            extract_clicked = st.button("🔍 EXTRACT BOQ DATA", use_container_width=True, type="primary")

        with col_manual:
            with st.expander("⚙️ Manual Column Override"):
                manual_cols = st.text_input("Column Names (comma separated)", 
                                            value=", ".join(st.session_state.column_structure or ["Item No", "Qty", "Fig No", "Description", "Material"]))
                if st.button("Apply Manual"):
                    st.session_state.column_structure = [c.strip() for c in manual_cols.split(",")]
                    st.success("Manual columns applied!")

        if extract_clicked:
            progress = st.progress(0, text="Initializing extraction...")

            # Use detected or manual column structure
            col_structure = st.session_state.column_structure or ["Item No", "Qty", "Fig No", "Description", "Material"]

            # Get material anchors
            material_anchors = st.session_state.get("material_anchors", [])

            progress.progress(30, text="Extracting data...")

            extractor = BOQExtractor(col_structure, material_anchors)
            items = extractor.extract_from_pdf(
                st.session_state.target_bytes,
                crop,
                max_items,
                page_range
            )

            progress.progress(100, text="Complete!")
            progress.empty()

            if items:
                st.session_state.data = pd.DataFrame(items)
                st.success(f"✅ Extracted {len(items)} items from PDF!")
                if material_anchors:
                    st.caption(f"🔧 Used {len(material_anchors)} material anchor(s) for detection")
            else:
                st.error("❌ No items found. Try adjusting crop area or check PDF format.")

    # TABLE DISPLAY
    if st.session_state.data is not None and not st.session_state.data.empty:
        df = st.session_state.data

        st.markdown("<div class='info-box'>Auto-detected columns with blank space handling | FIG NO preserved in description</div>", unsafe_allow_html=True)

        # Stats
        col_stats1, col_stats2, col_stats3 = st.columns(3)
        with col_stats1:
            st.metric("Total Items", len(df))
        with col_stats2:
            st.metric("Pages", df["Page"].nunique() if "Page" in df.columns else 1)
        with col_stats3:
            st.metric("With Material", len(df[df["Material"] != ""]) if "Material" in df.columns else 0)

        # Editable table
        edited_df = st.data_editor(
            df,
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            key="boq_table_v28",
            column_config={
                "Item": st.column_config.NumberColumn("Item", width="small"),
                "Qty": st.column_config.NumberColumn("Qty", width="small"),
                "Fig No": st.column_config.TextColumn("Fig No", width="medium"),
                "Description": st.column_config.TextColumn("Description", width="large"),
                "Material": st.column_config.TextColumn("Material", width="medium"),
                "Page": st.column_config.NumberColumn("Page", width="small")
            }
        )

        st.session_state.data = edited_df

        # Export
        st.divider()
        col_excel, col_csv, col_json = st.columns(3)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")

        with col_excel:
            excel = create_excel(edited_df, True)
            st.download_button("📥 Excel (.xlsx)", excel, f"BOQ_{ts}.xlsx",
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             use_container_width=True)

        with col_csv:
            csv = edited_df.to_csv(index=False).encode("utf-8")
            st.download_button("📄 CSV (.csv)", csv, f"BOQ_{ts}.csv", "text/csv", use_container_width=True)

        with col_json:
            json_data = edited_df.to_json(orient="records", indent=2)
            st.download_button("📋 JSON (.json)", json_data, f"BOQ_{ts}.json", "application/json", use_container_width=True)

    else:
        st.info("BOQ data will appear here after extraction. Upload PDF and click SCAN, then EXTRACT.")

    st.markdown("</div>", unsafe_allow_html=True)
